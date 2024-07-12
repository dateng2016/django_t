from django.http import HttpRequest
from django.shortcuts import render, redirect, HttpResponse
from openpyxl import load_workbook

from app02 import models
from app02.utils.pagination import Pagination
from app02.utils.form import *

# Create your views here.


def depart_list(request):
    queryset = models.Department.objects.all()
    return render(request, "depart_list.html", {"queryset": queryset})


def depart_add(request):
    if request.method == "GET":
        return render(request, "depart_add.html")
    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")


def depart_delete(request):
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(request, nid):
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, "depart_edit.html", {"row_object": row_object})
    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list/")


def depart_multi(request: HttpRequest):
    file_object = request.FILES.get("exc")
    # print(f"The type of the file is: {type(file_object)}")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # cell = sheet.cell(1, 1)
    # print(f"Here is the cell value: {cell.value}")
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart/list/")
